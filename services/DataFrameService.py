from entitys.general import materias_periodos,semestres,Ciclos,Areas
from openpyxl import load_workbook
import os
from typing import List, Dict, Optional,Union
import csv
from fastapi import  HTTPException


def process_excel_file(file_location: str):
    # Cargar el archivo Excel
    workbook = load_workbook(file_location)
    sheet = workbook['Sistemas']  # Acceder a la hoja 'Sistemas'

    # Borrar todas las hojas excepto 'Sistemas'
    sheet_names = workbook.sheetnames
    for sheet_name in sheet_names:
        if sheet_name != 'Sistemas':
            std = workbook[sheet_name]
            workbook.remove(std)

    # Insertar tres nuevas columnas al inicio
    sheet.insert_cols(1, 3)

    # Agregar los encabezados en la fila 3
    sheet.cell(row=3, column=1, value="Ciclo")
    sheet.cell(row=3, column=2, value="Semestre")
    sheet.cell(row=3, column=3, value="Area")

    # Llenar las columnas
    for row in range(4, sheet.max_row + 1):
        asignatura = sheet.cell(row=row, column=4).value  # La columna de asignaturas ahora es la 4
        if asignatura is None:
            continue

        # Inicializar variables de control
        ciclo_found = False
        semestre_found = False
        areas_found = False

        # Limpiar el nombre de la asignatura
        asignatura = str(asignatura).strip()

        # Buscar y asignar el ciclo
        for ciclo, asignaturas in Ciclos.items():
            if asignatura in asignaturas:
                sheet.cell(row=row, column=1, value=ciclo)
                ciclo_found = True
                break
        if not ciclo_found:
            sheet.cell(row=row, column=1, value="N/A")

        # Buscar y asignar el semestre
        for semestre, asignaturas in materias_periodos.items():
            if asignatura in asignaturas:
                sheet.cell(row=row, column=2, value=semestre)
                semestre_found = True
                break
        if not semestre_found:
            sheet.cell(row=row, column=2, value="N/A")

        # Buscar y asignar el área
        for area, asignaturas in Areas.items():
            if asignatura in asignaturas:
                sheet.cell(row=row, column=3, value=area)
                areas_found = True
                break
        if not areas_found:
            sheet.cell(row=row, column=3, value="N/A")

    # Modificar la fila 3 con los valores del vector semestre
    for col_index, semestre in enumerate(semestres, start=5):
        sheet.cell(row=3, column=col_index, value=semestre)

    # Guardar el archivo modificado
    workbook.save(file_location)

    # Convertir a CSV con formato específico
    csv_file_location = file_location.replace('.xlsx', '.csv')
    
    with open(csv_file_location, mode='w', newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quoting=csv.QUOTE_MINIMAL)
        
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            formatted_row = []
            for col_idx, cell in enumerate(row, start=1):
                # Manejar valores numéricos
                if col_idx >= 5 and row_idx >= 3:  # Datos numéricos desde la columna 4
                    try:
                        if cell is None or (isinstance(cell, str) and cell.strip() == ''):
                            formatted_value = '0'
                        elif isinstance(cell, (int, float)):
                            # Convertir a porcentaje y redondear a 3 decimales
                            value = float(cell)
                            if abs(value) < 1e-10:
                                formatted_value = '0'
                            else:
                                formatted_value = f"{value * 100:.3f}"
                        else:
                            # Intentar convertir strings que puedan ser números
                            value = float(str(cell).replace('%', '').strip())
                            formatted_value = f"{value:.3f}"
                    except (ValueError, TypeError):
                        formatted_value = '0'
                else:
                    # Valores no numéricos
                    formatted_value = str(cell) if cell is not None else ''
                
                formatted_row.append(formatted_value)
            
            # Escribir la fila solo si tiene contenido
            if any(cell.strip() != '' for cell in formatted_row):
                writer.writerow(formatted_row)
    
    # Borrar el archivo Excel original
    os.remove(file_location)

    return csv_file_location



def get_ciclo(asignatura: str):
    """
    Devuelve el ciclo correspondiente a una asignatura.
    """
    for ciclo, asignaturas in Ciclos.items():
        if asignatura in asignaturas:
            return ciclo
    return "N/A"

def get_asignaturas_por_ciclo(ciclo: str):
    """
    Devuelve las asignaturas correspondientes a un ciclo.
    """
    return Ciclos.get(ciclo, [])

def get_asignaturas_por_semestre(semestre: str):
    """
    Devuelve las asignaturas correspondientes a un ciclo.
    """
    return materias_periodos.get(semestre, [])


def get_data(materia: str, fecha_inicial: str, fecha_final: str, file_path: str) -> dict:
    try:
        # Validar fechas
        if fecha_inicial not in semestres or fecha_final not in semestres:
            raise ValueError("Las fechas deben estar en el rango permitido")
        
        if semestres.index(fecha_inicial) > semestres.index(fecha_final):
            raise ValueError("La fecha inicial debe ser anterior a la fecha final")
        
        # Leer el archivo CSV
        with open(file_path, 'r', encoding='utf-8') as file:
            csv_reader = csv.reader(file)
            
            # Saltar las dos primeras líneas de encabezado
            next(csv_reader)  # Saltar "EVOLUCION ASIGNATURAS CON MAYOR TASA DE MORTALIDAD"
            next(csv_reader)  # Saltar "INGENIERIA DE SISTEMAS"
            
            # Leer la línea de encabezados (tercera línea)
            headers = next(csv_reader)
            
            # Encontrar los índices de las columnas para el rango de fechas
            fechas_indices = []
            for i, header in enumerate(headers):
                if header in semestres:
                    if fecha_inicial <= header <= fecha_final:
                        fechas_indices.append(i)
            
            if not fechas_indices:
                raise ValueError("No se encontraron las fechas especificadas en el archivo")
            
            # Buscar la materia en el archivo
            for row in csv_reader:
                if len(row) < 3:  # Saltar filas vacías o mal formadas
                    continue
                    
                # La materia puede estar en diferentes posiciones según el formato
                asignatura_actual = None
                ciclo = row[0].strip()
                area = row[2].strip() if len(row) > 2 else ""
                
                # Extraer la asignatura del formato del archivo
                for campo in row:
                    if materia in campo:
                        asignatura_actual = materia
                        break
                
                if asignatura_actual == materia:
                    # Extraer los datos del rango solicitado
                    datos_rango = []
                    for idx in fechas_indices:
                        try:
                            if idx < len(row):
                                valor = row[idx].strip()
                                if valor == '' or valor == 'None':
                                    datos_rango.append(0.0)
                                else:
                                    # Eliminar el símbolo de porcentaje si existe y convertir a float
                                    valor = valor.replace('%', '').strip()
                                    datos_rango.append(float(valor))
                            else:
                                datos_rango.append(0.0)
                        except (ValueError, IndexError):
                            datos_rango.append(0.0)
                    
                    # Construir la respuesta
                    fechas_rango = [headers[i] for i in fechas_indices]
                    return {
                        "ciclo": ciclo,
                        "semestre": row[1] if len(row) > 1 else "semestre-desconocido",
                        "area": area,
                        "materia": materia,
                        "fechas": fechas_rango,
                        "datos": datos_rango
                    }
            
            # Si no se encuentra la materia
            raise ValueError(f"No se encontró la asignatura: {materia}")
            
    except Exception as e:
        raise ValueError(f"Error al procesar los datos: {str(e)}")
    



#nuevo metodo aun no probado
def get_all_subjects_data(file_path: str) -> list:
    try:
        materias_data = []
        
        # Leer el archivo CSV
        with open(file_path, 'r', encoding='utf-8') as file:
            csv_reader = csv.reader(file)
            
            # Saltar las dos primeras líneas de encabezado
            next(csv_reader)  # Saltar "EVOLUCION ASIGNATURAS CON MAYOR TASA DE MORTALIDAD"
            next(csv_reader)  # Saltar "INGENIERIA DE SISTEMAS"
            
            # Leer la línea de encabezados (tercera línea)
            headers = next(csv_reader)
            
            # Obtener los índices de todas las fechas (semestres)
            fechas_indices = []
            fechas_headers = []
            for i, header in enumerate(headers):
                if header in semestres:
                    fechas_indices.append(i)
                    fechas_headers.append(header)
            
            # Procesar cada fila del archivo
            for row in csv_reader:
                if len(row) < 3:  # Saltar filas vacías o mal formadas
                    continue
                
                # Extraer información básica
                ciclo = row[0].strip()
                semestre = row[1].strip() if len(row) > 1 else "semestre-desconocido"
                area = row[2].strip() if len(row) > 2 else ""
                
                # Encontrar el nombre de la materia
                materia = None
                for campo in row:
                    if campo.strip() and campo not in [ciclo, semestre, area]:
                        materia = campo.strip()
                        break
                
                if materia:
                    # Extraer todas las notas disponibles
                    datos_materia = []
                    for idx in fechas_indices:
                        try:
                            if idx < len(row):
                                valor = row[idx].strip()
                                if valor == '' or valor == 'None':
                                    datos_materia.append(0.0)
                                else:
                                    # Eliminar el símbolo de porcentaje si existe y convertir a float
                                    valor = valor.replace('%', '').strip()
                                    datos_materia.append(float(valor))
                            else:
                                datos_materia.append(0.0)
                        except (ValueError, IndexError):
                            datos_materia.append(0.0)
                    
                    # Construir el objeto de datos para esta materia
                    materia_obj = {
                        "ciclo": ciclo,
                        "semestre": semestre,
                        "area": area,
                        "materia": materia,
                        "fechas": fechas_headers,
                        "datos": datos_materia
                    }
                    
                    materias_data.append(materia_obj)
            
            if not materias_data:
                raise ValueError("No se encontraron materias en el archivo")
            
            return materias_data
            
    except Exception as e:
        raise ValueError(f"Error al procesar los datos: {str(e)}")

#fin metodo
def get_tasa_mortandad(ciclo: str) -> list:
    try:
        # Listar archivos en el directorio 'uploads/'
        directory = 'uploads/'
        files = os.listdir(directory)
        if not files:
            raise FileNotFoundError("No se encontró ningún archivo en el directorio.")

        # Asumir que hay solo un archivo en el directorio
        file_path = os.path.join(directory, files[0])

        # Leer el archivo CSV
        with open(file_path, 'r', encoding='utf-8') as file:
            csv_reader = csv.reader(file)
            
            # Saltar las dos primeras líneas de encabezado
            next(csv_reader)  # Saltar "EVOLUCION ASIGNATURAS CON MAYOR TASA DE MORTALIDAD"
            next(csv_reader)  # Saltar "INGENIERIA DE SISTEMAS"
            
            # Filtrar las filas que contienen el ciclo específico
            filtered_rows = [row for row in csv_reader if ciclo in row]
            
            if not filtered_rows:
                return None
            
            # Convertir las filas filtradas a una lista de diccionarios
            keys = ["Ciclo", "Semestre", "Area", "Asignatura", "2017-1", "2017-2", "2018-1","2018-2", "2019-1", "2019-2", "2020-1", "2020-2","2021-1", "2021-2", "2022-1", "2022-2", "2023-1", "2023-2", "2024-1"]
            result = [
                {k: v for k, v in zip(keys, row) if k not in {"Ciclo", "Semestre", "Area"}}
                for row in filtered_rows
            ]
            
            return result
    
    except FileNotFoundError:
        raise FileNotFoundError("Archivo no encontrado.")
    except Exception as e:
        raise Exception(f"Error al procesar el archivo: {str(e)}")

#areas

def get_asignaturas_por_area(area: str):
    """
    Devuelve las asignaturas correspondientes a un ciclo.
    """
    return Areas.get(area, [])


def get_tasa_mortandad_areas(areas: str) -> list:
    try:
        # Listar archivos en el directorio 'uploads/'
        directory = 'uploads/'
        files = os.listdir(directory)
        if not files:
            raise FileNotFoundError("No se encontró ningún archivo en el directorio.")

        # Asumir que hay solo un archivo en el directorio
        file_path = os.path.join(directory, files[0])

        # Leer el archivo CSV
        with open(file_path, 'r', encoding='utf-8') as file:
            csv_reader = csv.reader(file)
            
            # Saltar las dos primeras líneas de encabezado
            next(csv_reader)  # Saltar "EVOLUCION ASIGNATURAS CON MAYOR TASA DE MORTALIDAD"
            next(csv_reader)  # Saltar "INGENIERIA DE SISTEMAS"
            
            # Filtrar las filas que contienen el ciclo específico
            filtered_rows = [row for row in csv_reader if areas in row]
            
            if not filtered_rows:
                return None
            
            # Convertir las filas filtradas a una lista de diccionarios
            keys = ["Ciclo", "Semestre", "Area", "Asignatura", "2017-1", "2017-2", "2018-1","2018-2", "2019-1", "2019-2", "2020-1", "2020-2","2021-1", "2021-2", "2022-1", "2022-2", "2023-1", "2023-2", "2024-1"]
            result = [
                {k: v for k, v in zip(keys, row) if k not in {"Ciclo", "Semestre", "Area"}}
                for row in filtered_rows
            ]
            
            return result
    
    except FileNotFoundError:
        raise FileNotFoundError("Archivo no encontrado.")
    except Exception as e:
        raise Exception(f"Error al procesar el archivo: {str(e)}")


#Generales

def cargar_datos_csv(directorio: str) -> list:
        try:
            # Verificar que el directorio existe
            if not os.path.exists(directorio):
                raise HTTPException(
                    status_code=404,
                    detail="El directorio no existe"
                )

            # Obtener la lista de archivos en el directorio
            archivos = [f for f in os.listdir(directorio) if os.path.isfile(os.path.join(directorio, f))]
            
            # Filtrar solo los archivos CSV
            archivos_csv = [f for f in archivos if f.lower().endswith('.csv')]

            if not archivos_csv:
                raise HTTPException(
                    status_code=404,
                    detail="No se encontró ningún archivo CSV en el directorio especificado."
                )

            # Tomar el primer archivo CSV
            archivo_csv = archivos_csv[0]
            ruta_csv = os.path.join(directorio, archivo_csv)

            # Leer el archivo CSV
            with open(ruta_csv, 'r', encoding='utf-8') as file:
                csv_reader = csv.reader(file)
                data = [row for row in csv_reader]
                
                if not data:
                    raise HTTPException(
                        status_code=404,
                        detail="El archivo CSV está vacío"
                    )
                    
                return data

        except Exception as e:
            if isinstance(e, HTTPException):
                raise e
            raise HTTPException(
                status_code=500,
                detail=f"Error al cargar el archivo CSV: {str(e)}"
            )

def obtener_tasas_mortalidad_mas_alta(data: list) -> list:
        try:
            # Los encabezados reales están en la línea 3 (índice 2)
            header = data[2]
            
            # Lista de semestres a buscar
            semestres = [
                "2017-1", "2017-2", "2018-1", "2018-2", 
                "2019-1", "2019-2", "2020-1", "2020-2", 
                "2021-1", "2021-2", "2022-1", "2022-2", 
                "2023-1", "2023-2", "2024-1"
            ]
            
            # Encontrar índices de las columnas de semestres
            indices_semestres = []
            for semestre in semestres:
                try:
                    idx = header.index(semestre)
                    indices_semestres.append((semestre, idx))
                except ValueError:
                    continue

            if not indices_semestres:
                raise HTTPException(
                    status_code=404,
                    detail="No se encontraron columnas de semestres válidos en el CSV"
                )

            # Índices de las columnas importantes
            try:
                idx_ciclo = header.index("Ciclo")
                idx_asignatura = header.index("ASIGNATURA")
            except ValueError:
                raise HTTPException(
                    status_code=404,
                    detail="No se encontraron las columnas requeridas (Ciclo, ASIGNATURA)"
                )

            resultados = []
            # Analizar cada semestre
            for semestre, idx in indices_semestres:
                max_tasa = -float('inf')
                asignatura_max = ""
                ciclo_max = ""
                
                # Comenzar desde la línea 4 (índice 3) para saltar los encabezados
                for row in data[3:]:
                    try:
                        if len(row) > idx and row[idx] and row[idx].strip():
                            tasa = float(row[idx])
                            if tasa > max_tasa:
                                max_tasa = tasa
                                asignatura_max = row[idx_asignatura]
                                ciclo_max = row[idx_ciclo]
                    except (ValueError, IndexError):
                        continue
                
                if asignatura_max and max_tasa > -float('inf'):
                    resultado = {
                        "semestre": semestre,
                        "ciclo": ciclo_max,
                        "asignatura": asignatura_max,
                        "tasa_mortalidad": round(max_tasa, 3)
                    }
                    resultados.append(resultado)
            
            if not resultados:
                raise HTTPException(
                    status_code=404,
                    detail="No se encontraron tasas de mortalidad válidas"
                )

            return resultados

        except Exception as e:
            if isinstance(e, HTTPException):
                raise e
            raise HTTPException(
                status_code=500,
                detail=f"Error en el análisis de datos: {str(e)}"
            )


def cargar_y_analizar_datos(directorio: str) -> List[Dict[str, Union[str, float]]]:
        data = cargar_datos_csv(directorio)
        return obtener_tasas_mortalidad_mas_alta(data)